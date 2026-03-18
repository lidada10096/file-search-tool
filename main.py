#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能文件搜索工具 - 命令行版本

通用性设计：支持任意关键词在指定目录下匹配关键文件

使用方法:
  1. python main.py                    # 基础搜索（仅用项目名称）
  2. python main.py 关键词1 关键词2    # 关键词增加匹配权重
  3. python main.py run                # 执行复制（按确认表）

评分规则:
  - 项目名称匹配度: 0-100分（主要依据）
  - 文件名匹配关键词: +10分
  - 路径匹配关键词: +10分
  - 文件名和路径同时匹配: +15分（额外奖励）
  - 总分 = 名称匹配度 + 关键词加分

作者：AI Assistant
日期：2026-03-18
"""

import os
import sys
import json
import shutil
import re
from datetime import datetime
from typing import List, Dict, Tuple
from difflib import SequenceMatcher

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("请先安装openpyxl: pip install openpyxl")
    sys.exit(1)


class ScoringMatcher:
    """评分匹配器 - 综合考虑项目名称和关键词"""
    
    def __init__(self, threshold: float = 0.8):
        self.threshold = threshold
    
    def similarity(self, s1: str, s2: str) -> float:
        """计算字符串相似度 (0-1)"""
        if not s1 or not s2:
            return 0.0
        return SequenceMatcher(None, s1.lower(), s2.lower()).ratio()
    
    def calculate_score(self, item_name: str, filepath: str, 
                        keywords: List[str]) -> Tuple[float, Dict]:
        """
        计算文件匹配评分 - 关键词同时检查文件名和路径，两者都匹配加分更多
        
        Returns:
            (总分, 评分详情)
        """
        filename = os.path.basename(filepath)
        filename_no_ext = os.path.splitext(filename)[0]
        
        # 1. 项目名称匹配度 (0-100分，权重最高)
        item_sim = self.similarity(item_name, filename_no_ext)
        
        # 如果项目名完全包含在文件名中，提高基础分
        if item_name.lower() in filename_no_ext.lower():
            item_sim = max(item_sim, 0.85)
        
        item_score = item_sim * 100  # 转换为百分制
        
        # 2. 关键词加分 - 优化：文件名和路径同时匹配加分更多
        keyword_bonus = 0
        matched_keywords = []
        
        for kw in keywords:
            kw_lower = kw.lower()
            matched_in = []  # 记录关键词匹配的位置
            
            # 检查文件名
            if kw_lower in filename.lower():
                matched_in.append("文件名")
            
            # 检查路径（不包含文件名部分）
            dir_path = os.path.dirname(filepath).lower()
            if kw_lower in dir_path:
                matched_in.append("路径")
            
            # 根据匹配位置数量加分
            if len(matched_in) == 2:
                # 文件名和路径都匹配 +15分
                keyword_bonus += 15
                matched_keywords.append(f"{kw}(文件名+路径)")
            elif len(matched_in) == 1:
                # 只匹配一处 +10分
                keyword_bonus += 10
                matched_keywords.append(f"{kw}({matched_in[0]})")
        
        keyword_bonus = min(keyword_bonus, 45)  # 最高45分（3个关键词都同时匹配文件名和路径）
        
        # 3. 总分
        total_score = item_score + keyword_bonus
        
        details = {
            'item_score': round(item_score, 1),
            'keyword_bonus': keyword_bonus,
            'total_score': round(total_score, 1),
            'matched_keywords': matched_keywords,
            'item_sim': item_sim
        }
        
        return total_score, details
    
    def is_potential_match(self, item_name: str, filepath: str) -> bool:
        """判断是否为潜在匹配（宽松条件，确保不遗漏）"""
        filename = os.path.basename(filepath)
        filename_no_ext = os.path.splitext(filename)[0]
        
        # 条件1: 项目名称与文件名相似度 >= 阈值
        sim = self.similarity(item_name, filename_no_ext)
        if sim >= self.threshold:
            return True
        
        # 条件2: 项目名称完全包含在文件名中
        if item_name.lower() in filename_no_ext.lower():
            return True
        
        # 条件3: 文件名完全包含在项目名称中
        if filename_no_ext.lower() in item_name.lower():
            return True
        
        return False


class ItemNameParser:
    """项目名称解析器"""
    
    @staticmethod
    def parse(filepath: str) -> List[str]:
        items = []
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                line = re.sub(r'^\d+\s*[→\.\)\-\s]+', '', line)
                line = line.strip()
                if line:
                    items.append(line)
        return items


def generate_excel(results: List[Dict], item_names: List[str], output_path: str, keywords: List[str]):
    """生成搜索结果确认表 - 按项目分组，组内按评分排序，包含未匹配的项目"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "搜索结果"
    
    # 标题
    ws['A1'] = "文件搜索结果（按项目分组，组内按评分排序）"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:K1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 副标题
    keyword_info = f" | 关键词: {', '.join(keywords)}" if keywords else ""
    total_items = len(item_names)
    matched_items = len(set(item['item'] for item in results))
    unmatched_count = total_items - matched_items
    ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}{keyword_info} | 项目总数: {total_items} | 匹配: {matched_items} | 未匹配: {unmatched_count}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:K2')
    
    # 使用说明
    ws['A3'] = '使用说明: 1) 灰色行表示未找到匹配文件的项目；2) 同一项目的多匹配项用颜色区分；3) 建议只选择每组中评分最高的1个文件'
    ws['A3'].font = Font(size=9, color="FF0000")
    ws.merge_cells('A3:K3')
    
    # 表头
    headers = ['序号', '项目名称', '匹配数', '源文件路径', '源文件名', '目标文件名', 
               '名称匹配度', '关键词加分', '总分', '是否复制', '备注']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 按项目名称分组
    from collections import defaultdict
    item_groups = defaultdict(list)
    for item in results:
        item_groups[item['item']].append(item)
    
    # 找出未匹配的项目
    matched_item_names = set(item_groups.keys())
    unmatched_items = [name for name in item_names if name not in matched_item_names]
    
    # 统计有多个匹配的项目
    multi_match_items = {item: items for item, items in item_groups.items() if len(items) > 1}
    
    # 定义颜色方案 - 每组不同颜色
    group_colors = [
        PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"),  # 浅蓝
        PatternFill(start_color="F3E8FD", end_color="F3E8FD", fill_type="solid"),  # 浅紫
        PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),  # 浅红
        PatternFill(start_color="FDF3E8", end_color="FDF3E8", fill_type="solid"),  # 浅橙
        PatternFill(start_color="FDFDE8", end_color="FDFDE8", fill_type="solid"),  # 浅黄
        PatternFill(start_color="E8FDE8", end_color="E8FDE8", fill_type="solid"),  # 浅绿
    ]
    
    # 未匹配项目的灰色背景
    unmatched_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    unmatched_font = Font(italic=True, color="666666")
    
    # 高亮重复项的背景色
    duplicate_header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    duplicate_header_font = Font(bold=True, color="FFFFFF")
    
    # 填充数据 - 按项目分组，每组内按评分降序
    row = 6
    idx = 1
    color_idx = 0
    
    # 先显示有匹配的项目（按原始顺序）
    for item_name in item_names:
        if item_name not in item_groups:
            continue
            
        items = item_groups[item_name]
        # 组内按评分降序排序
        items.sort(key=lambda x: x['total_score'], reverse=True)
        
        match_count = len(items)
        is_duplicate = match_count > 1
        
        # 为当前组选择颜色
        group_color = group_colors[color_idx % len(group_colors)]
        color_idx += 1
        
        for item_idx, item in enumerate(items):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item['item'])
            ws.cell(row=row, column=3, value=match_count)  # 匹配数
            ws.cell(row=row, column=4, value=item['path'])
            ws.cell(row=row, column=5, value=item['filename'])
            ws.cell(row=row, column=6, value=item['filename'])
            ws.cell(row=row, column=7, value=f"{item['item_score']:.1f}")
            ws.cell(row=row, column=8, value=item['keyword_bonus'])
            ws.cell(row=row, column=9, value=f"{item['total_score']:.1f}")
            # 默认只选中每组第一个（评分最高的）
            ws.cell(row=row, column=10, value=(item_idx == 0))
            ws.cell(row=row, column=11, value=", ".join(item.get('matched_keywords', [])))
            
            # 整行背景色（分组颜色）
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = group_color
            
            # 如果是多匹配项，项目名称和匹配数列高亮显示
            if is_duplicate:
                ws.cell(row=row, column=2).fill = duplicate_header_fill
                ws.cell(row=row, column=2).font = duplicate_header_font
                ws.cell(row=row, column=3).fill = duplicate_header_fill
                ws.cell(row=row, column=3).font = duplicate_header_font
            
            # 总分颜色标识（覆盖背景色）
            score_cell = ws.cell(row=row, column=9)
            if item['total_score'] >= 90:
                score_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                score_cell.font = Font(bold=True)
            elif item['total_score'] >= 70:
                score_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            # 居中对齐
            for col in [1, 3, 7, 8, 9, 10]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
            idx += 1
    
    # 显示未匹配的项目（灰色背景）
    if unmatched_items:
        for item_name in unmatched_items:
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item_name)
            ws.cell(row=row, column=3, value=0)  # 匹配数为0
            ws.cell(row=row, column=4, value="未找到匹配文件")
            ws.cell(row=row, column=5, value="-")
            ws.cell(row=row, column=6, value="-")
            ws.cell(row=row, column=7, value="-")
            ws.cell(row=row, column=8, value="-")
            ws.cell(row=row, column=9, value="-")
            ws.cell(row=row, column=10, value="FALSE")
            ws.cell(row=row, column=11, value="未匹配")
            
            # 整行灰色背景
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.fill = unmatched_fill
                cell.font = unmatched_font
            
            # 项目名称列用红色背景突出显示未匹配
            ws.cell(row=row, column=2).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=3).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws.cell(row=row, column=3).font = Font(bold=True, color="FFFFFF")
            
            # 居中对齐
            for col in [1, 3, 10]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
            idx += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 30
    
    ws.freeze_panes = 'A6'
    wb.save(output_path)
    
    # 输出统计信息
    print(f"\n统计信息:")
    print(f"  项目总数: {total_items}")
    print(f"  匹配到文件的项目: {matched_items}")
    print(f"  未匹配的项目: {len(unmatched_items)}")
    print(f"  总匹配文件数: {len(results)}")
    if multi_match_items:
        print(f"  有多个匹配的项目: {len(multi_match_items)} 个")
        for item, items in sorted(multi_match_items.items(), key=lambda x: len(x[1]), reverse=True)[:5]:
            print(f"    - {item}: {len(items)} 个匹配")
    if unmatched_items:
        print(f"  未匹配项目列表（前10个）:")
        for item in unmatched_items[:10]:
            print(f"    - {item}")
        if len(unmatched_items) > 10:
            print(f"    ... 还有 {len(unmatched_items) - 10} 个未匹配项目")
    
    return output_path


def execute_copy(excel_path: str, target_dir: str):
    """根据Excel执行复制"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    os.makedirs(target_dir, exist_ok=True)
    
    tasks = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row[0]:
            continue
        # 新的列顺序: 序号, 项目名称, 匹配数, 源文件路径, 源文件名, 目标文件名, 
        #            名称匹配度, 关键词加分, 总分, 是否复制, 备注
        rank, item, match_count, source, src_name, tgt_name, item_score, kw_bonus, total, should_copy, remark = row[:11]
        if should_copy == True or (isinstance(should_copy, str) and str(should_copy).upper() in ['TRUE', '是', 'Y', 'YES', '1']):
            if source and os.path.exists(source):
                tasks.append({'source': source, 'target': tgt_name or src_name})
    
    print(f"共 {len(tasks)} 个文件待复制")
    
    copied = failed = 0
    for i, task in enumerate(tasks, 1):
        source = task['source']
        filename = task['target']
        
        # 处理重名
        target_path = os.path.join(target_dir, filename)
        if os.path.exists(target_path):
            name, ext = os.path.splitext(filename)
            counter = 1
            while os.path.exists(target_path):
                target_path = os.path.join(target_dir, f"{name}_{counter:03d}{ext}")
                counter += 1
        
        try:
            shutil.copy2(source, target_path)
            copied += 1
            print(f"  [{i}/{len(tasks)}] ✓ {os.path.basename(target_path)}")
        except Exception as e:
            failed += 1
            print(f"  [{i}/{len(tasks)}] ✗ {filename} - {e}")
    
    return copied, failed


def find_latest_excel(target_dir: str) -> str:
    """查找最新的搜索结果表"""
    if not os.path.exists(target_dir):
        return None
    
    excels = [f for f in os.listdir(target_dir) if f.startswith('搜索结果_') and f.endswith('.xlsx')]
    if not excels:
        return None
    
    excels.sort(reverse=True)
    return os.path.join(target_dir, excels[0])


def main():
    # 加载配置
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    items_file = config.get('items_file', 'name.txt')
    source_dir = config.get('source_dir', './drawings')
    target_dir = config.get('target_dir', './output')
    
    # 解析命令行参数
    args = sys.argv[1:]
    
    if len(args) == 1 and args[0].lower() == 'run':
        # 执行复制模式
        print("=" * 70)
        print("执行复制")
        print("=" * 70)
        
        excel_path = find_latest_excel(target_dir)
        if not excel_path:
            print(f"\n✗ 错误: 在 {target_dir} 中找不到搜索结果表")
            print("\n请先运行搜索:")
            print("  python main.py")
            print("或带关键词:")
            print("  python main.py 关键词1 关键词2")
            sys.exit(1)
        
        print(f"✓ 找到搜索结果表: {os.path.basename(excel_path)}")
        copied, failed = execute_copy(excel_path, target_dir)
        
        print(f"\n复制完成: 成功 {copied}, 失败 {failed}")
        print(f"文件已保存到: {target_dir}")
        return
    
    # 搜索模式
    keywords = args  # 所有参数都作为关键词
    
    print("=" * 70)
    print("智能文件搜索工具（评分版）")
    print("=" * 70)
    
    if keywords:
        print(f"关键词: {', '.join(keywords)}")
    else:
        print("关键词: 无（仅用项目名称匹配）")
    
    # 读取项目名称
    print(f"\n读取项目名称: {items_file}")
    item_names = ItemNameParser.parse(items_file)
    print(f"共 {len(item_names)} 个项目")
    
    # 扫描所有文件
    print(f"\n扫描源目录: {source_dir}")
    all_files = []
    for root, dirs, files in os.walk(source_dir):
        for filename in files:
            all_files.append(os.path.join(root, filename))
    print(f"共发现 {len(all_files)} 个文件")
    
    # 评分匹配
    print(f"\n开始评分匹配...")
    matcher = ScoringMatcher(threshold=0.8)  # 阈值0.8确保相关性
    
    all_results = []
    
    for item_idx, item_name in enumerate(item_names, 1):
        print(f"  匹配 [{item_idx}/{len(item_names)}]: {item_name}")
        
        for filepath in all_files:
            # 检查是否为潜在匹配
            if matcher.is_potential_match(item_name, filepath):
                # 计算评分
                total_score, details = matcher.calculate_score(item_name, filepath, keywords)
                
                all_results.append({
                    'item': item_name,
                    'path': filepath,
                    'filename': os.path.basename(filepath),
                    'item_score': details['item_score'],
                    'keyword_bonus': details['keyword_bonus'],
                    'total_score': details['total_score'],
                    'matched_keywords': details['matched_keywords']
                })
    
    # 按总分排序
    all_results.sort(key=lambda x: x['total_score'], reverse=True)
    
    print(f"\n搜索完成，共找到 {len(all_results)} 个潜在匹配")
    
    if all_results:
        # 显示前10个结果
        print("\n评分最高的前10个结果:")
        print("-" * 70)
        for i, item in enumerate(all_results[:10], 1):
            print(f"{i}. {item['filename']}")
            print(f"   项目: {item['item']}")
            print(f"   评分: 名称{item['item_score']:.1f} + 关键词{item['keyword_bonus']} = 总分{item['total_score']:.1f}")
            if item['matched_keywords']:
                print(f"   关键词: {', '.join(item['matched_keywords'])}")
            print()
        
        # 生成Excel
        os.makedirs(target_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = os.path.join(target_dir, f"搜索结果_{timestamp}.xlsx")
        generate_excel(all_results, item_names, excel_path, keywords)
        
        print("=" * 70)
        print(f"✓ 搜索结果表已生成: {excel_path}")
        print("=" * 70)
        print("\n请打开Excel查看所有结果（已按评分排序）")
        print("在'是否复制'列选择TRUE/FALSE，确认后运行:")
        print("  python main.py run")
    else:
        print("\n未找到任何匹配文件")
        print("建议:")
        print("  1. 检查源目录路径是否正确")
        print("  2. 尝试添加关键词扩大搜索范围")
        print("  3. 检查项目名称文件是否正确")


if __name__ == '__main__':
    main()
