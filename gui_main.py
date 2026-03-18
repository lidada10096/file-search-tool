#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能文件搜索工具 - GUI版本

图形界面版本，支持：
1. 搜索配置设置
2. 执行搜索并显示进度
3. 查看搜索结果
4. 选择文件并执行复制

通用性设计：支持任意关键词在指定目录下匹配关键文件
"""

import os
import sys
import json
import shutil
import re
import threading
from datetime import datetime
from typing import List, Dict, Tuple
from difflib import SequenceMatcher
from collections import defaultdict

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog, scrolledtext
except ImportError:
    print("错误: 需要tkinter库")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("错误: 需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


class ScoringMatcher:
    """评分匹配器"""
    
    def __init__(self, threshold: float = 0.8):
        self.threshold = threshold
    
    def similarity(self, s1: str, s2: str) -> float:
        if not s1 or not s2:
            return 0.0
        return SequenceMatcher(None, s1.lower(), s2.lower()).ratio()
    
    def is_match(self, name: str, filename: str) -> Tuple[bool, float]:
        filename_no_ext = os.path.splitext(filename)[0]
        sim = self.similarity(name, filename_no_ext)
        name_lower = name.lower()
        filename_lower = filename_no_ext.lower()
        if name_lower in filename_lower or filename_lower in name_lower:
            sim = max(sim, 0.85)
        return sim >= self.threshold, sim
    
    def calculate_score(self, item_name: str, filepath: str, keywords: List[str]) -> Tuple[float, Dict]:
        """计算文件匹配评分 - 关键词同时检查文件名和路径，两者都匹配加分更多"""
        filename = os.path.basename(filepath)
        filename_no_ext = os.path.splitext(filename)[0]
        
        # 项目名称匹配度
        item_sim = self.similarity(item_name, filename_no_ext)
        if item_name.lower() in filename_no_ext.lower():
            item_sim = max(item_sim, 0.85)
        item_score = item_sim * 100
        
        # 关键词加分 - 优化：文件名和路径同时匹配加分更多
        keyword_bonus = 0
        matched_keywords = []
        
        for kw in keywords:
            kw_lower = kw.lower()
            matched_in = []  # 记录关键词匹配的位置
            
            # 检查文件名
            if kw_lower in filename.lower():
                matched_in.append("文件名")
            
            # 检查路径（不包含文件名部分）
            dir_path = os.path.dirname(filepath).lower()
            if kw_lower in dir_path:
                matched_in.append("路径")
            
            # 根据匹配位置数量加分
            if len(matched_in) == 2:
                # 文件名和路径都匹配 +15分
                keyword_bonus += 15
                matched_keywords.append(f"{kw}(文件名+路径)")
            elif len(matched_in) == 1:
                # 只匹配一处 +10分
                keyword_bonus += 10
                matched_keywords.append(f"{kw}({matched_in[0]})")
        
        keyword_bonus = min(keyword_bonus, 45)  # 最高45分（3个关键词都同时匹配文件名和路径）
        total_score = item_score + keyword_bonus
        
        return total_score, {
            'item_score': round(item_score, 1),
            'keyword_bonus': keyword_bonus,
            'total_score': round(total_score, 1),
            'matched_keywords': matched_keywords
        }
    
    def is_potential_match(self, item_name: str, filepath: str) -> bool:
        """判断是否为潜在匹配"""
        filename = os.path.basename(filepath)
        filename_no_ext = os.path.splitext(filename)[0]
        
        sim = self.similarity(item_name, filename_no_ext)
        if sim >= self.threshold:
            return True
        if item_name.lower() in filename_no_ext.lower():
            return True
        if filename_no_ext.lower() in item_name.lower():
            return True
        return False


class ItemNameParser:
    """项目名称解析器"""
    
    @staticmethod
    def parse(filepath: str) -> List[str]:
        """解析项目名称列表文件"""
        items = []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    # 移除行号前缀
                    line = re.sub(r'^\d+\s*[→\.\)\-\s]+', '', line)
                    line = line.strip()
                    if line:
                        items.append(line)
        except Exception as e:
            print(f"读取项目文件错误: {e}")
        return items


class SearchApp:
    """搜索工具主界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("智能文件搜索工具 v2.0")
        self.root.geometry("900x700")
        
        # 数据
        self.config = {}
        self.search_results = []
        self.load_config()
        
        # 创建界面
        self.create_widgets()
    
    def load_config(self):
        """加载配置文件"""
        try:
            with open('config.json', 'r', encoding='utf-8') as f:
                self.config = json.load(f)
        except:
            self.config = {
                'items_file': 'name.txt',
                'source_dir': '',
                'target_dir': '',
                'threshold': 0.8
            }
    
    def save_config(self):
        """保存配置文件"""
        try:
            with open('config.json', 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败: {e}")
    
    def create_widgets(self):
        """创建界面组件"""
        # 标题
        title = tk.Label(self.root, text="智能文件搜索工具", font=("微软雅黑", 20, "bold"))
        title.pack(pady=10)
        
        # 配置区域
        config_frame = tk.LabelFrame(self.root, text="配置", font=("微软雅黑", 10))
        config_frame.pack(fill=tk.X, padx=20, pady=5)
        
        # 项目名称文件
        tk.Label(config_frame, text="项目名称文件:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.items_entry = tk.Entry(config_frame, width=50)
        self.items_entry.grid(row=0, column=1, padx=5, pady=5)
        self.items_entry.insert(0, self.config.get('items_file', 'name.txt'))
        tk.Button(config_frame, text="浏览...", command=self.browse_items).grid(row=0, column=2, padx=5)
        
        # 源目录
        tk.Label(config_frame, text="源目录:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.source_entry = tk.Entry(config_frame, width=50)
        self.source_entry.grid(row=1, column=1, padx=5, pady=5)
        self.source_entry.insert(0, self.config.get('source_dir', ''))
        tk.Button(config_frame, text="浏览...", command=self.browse_source).grid(row=1, column=2, padx=5)
        
        # 目标目录
        tk.Label(config_frame, text="目标目录:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.target_entry = tk.Entry(config_frame, width=50)
        self.target_entry.grid(row=2, column=1, padx=5, pady=5)
        self.target_entry.insert(0, self.config.get('target_dir', ''))
        tk.Button(config_frame, text="浏览...", command=self.browse_target).grid(row=2, column=2, padx=5)
        
        # 关键词
        tk.Label(config_frame, text="关键词(可选):").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.keywords_entry = tk.Entry(config_frame, width=50)
        self.keywords_entry.grid(row=3, column=1, padx=5, pady=5)
        tk.Label(config_frame, text="多个关键词用空格分隔").grid(row=3, column=2, sticky=tk.W)
        
        # 保存配置按钮
        tk.Button(config_frame, text="保存配置", command=self.save_current_config).grid(row=4, column=1, pady=10)
        
        # 操作按钮区域
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="开始搜索", font=("微软雅黑", 12), 
                 bg="#4472C4", fg="white", width=15, command=self.start_search).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="执行复制", font=("微软雅黑", 12),
                 bg="#70AD47", fg="white", width=15, command=self.start_copy).pack(side=tk.LEFT, padx=10)
        
        # 进度条
        self.progress = ttk.Progressbar(self.root, length=800, mode='determinate')
        self.progress.pack(pady=10)
        
        self.status_label = tk.Label(self.root, text="就绪", font=("微软雅黑", 10))
        self.status_label.pack()
        
        # 结果显示区域
        result_frame = tk.LabelFrame(self.root, text="搜索结果", font=("微软雅黑", 10))
        result_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        # 创建表格
        columns = ('项目名称', '文件名', '名称匹配度', '关键词加分', '总分', '是否复制')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=15)
        
        # 设置列宽
        self.tree.column('项目名称', width=200)
        self.tree.column('文件名', width=300)
        self.tree.column('名称匹配度', width=80, anchor='center')
        self.tree.column('关键词加分', width=80, anchor='center')
        self.tree.column('总分', width=60, anchor='center')
        self.tree.column('是否复制', width=80, anchor='center')
        
        # 设置列标题
        for col in columns:
            self.tree.heading(col, text=col)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 双击切换是否复制
        self.tree.bind('<Double-1>', self.toggle_copy)
        
        # 底部按钮
        bottom_frame = tk.Frame(self.root)
        bottom_frame.pack(pady=10)
        
        tk.Button(bottom_frame, text="全选", command=self.select_all).pack(side=tk.LEFT, padx=5)
        tk.Button(bottom_frame, text="全不选", command=self.select_none).pack(side=tk.LEFT, padx=5)
        tk.Button(bottom_frame, text="仅选每组最高分", command=self.select_best).pack(side=tk.LEFT, padx=5)
        tk.Button(bottom_frame, text="导出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
    
    def browse_items(self):
        filename = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if filename:
            self.items_entry.delete(0, tk.END)
            self.items_entry.insert(0, filename)
    
    def browse_source(self):
        directory = filedialog.askdirectory()
        if directory:
            self.source_entry.delete(0, tk.END)
            self.source_entry.insert(0, directory)
    
    def browse_target(self):
        directory = filedialog.askdirectory()
        if directory:
            self.target_entry.delete(0, tk.END)
            self.target_entry.insert(0, directory)
    
    def save_current_config(self):
        self.config['items_file'] = self.items_entry.get()
        self.config['source_dir'] = self.source_entry.get()
        self.config['target_dir'] = self.target_entry.get()
        self.save_config()
        messagebox.showinfo("成功", "配置已保存")
    
    def start_search(self):
        """开始搜索"""
        items_file = self.items_entry.get()
        source_dir = self.source_entry.get()
        keywords = self.keywords_entry.get().split()
        
        if not os.path.exists(items_file):
            messagebox.showerror("错误", "项目名称文件不存在")
            return
        if not os.path.exists(source_dir):
            messagebox.showerror("错误", "源目录不存在")
            return
        
        # 在新线程中执行搜索
        thread = threading.Thread(target=self.do_search, args=(items_file, source_dir, keywords))
        thread.daemon = True
        thread.start()
    
    def do_search(self, items_file, source_dir, keywords):
        """执行搜索"""
        self.status_label.config(text="正在读取项目名称...")
        
        item_names = ItemNameParser.parse(items_file)
        if not item_names:
            self.root.after(0, lambda: messagebox.showerror("错误", "未能读取到项目名称"))
            return
        
        self.item_names = item_names  # 保存所有项目名称
        self.root.after(0, lambda: self.status_label.config(text=f"读取了 {len(item_names)} 个项目名称"))
        
        # 扫描文件
        self.root.after(0, lambda: self.status_label.config(text="正在扫描文件..."))
        all_files = []
        for root, dirs, files in os.walk(source_dir):
            for filename in files:
                all_files.append(os.path.join(root, filename))
        
        self.root.after(0, lambda: self.status_label.config(text=f"扫描到 {len(all_files)} 个文件，开始匹配..."))
        
        # 匹配
        matcher = ScoringMatcher(threshold=0.8)
        results = []
        
        for idx, item_name in enumerate(item_names):
            progress = (idx + 1) / len(item_names) * 100
            self.root.after(0, lambda p=progress: self.progress.config(value=p))
            self.root.after(0, lambda s=item_name: self.status_label.config(text=f"正在匹配: {s}"))
            
            for filepath in all_files:
                if matcher.is_potential_match(item_name, filepath):
                    total_score, details = matcher.calculate_score(item_name, filepath, keywords)
                    results.append({
                        'item': item_name,
                        'path': filepath,
                        'filename': os.path.basename(filepath),
                        'item_score': details['item_score'],
                        'keyword_bonus': details['keyword_bonus'],
                        'total_score': details['total_score'],
                        'matched_keywords': details['matched_keywords'],
                        'selected': True,  # 默认选中
                        'matched': True    # 有匹配
                    })
        
        # 按项目分组，组内按评分排序
        item_groups = defaultdict(list)
        for item in results:
            item_groups[item['item']].append(item)
        
        # 找出未匹配的项目
        matched_item_names = set(item_groups.keys())
        unmatched_items = [name for name in item_names if name not in matched_item_names]
        
        sorted_results = []
        # 先添加有匹配的项目（按原始顺序）
        for item_name in item_names:
            if item_name in item_groups:
                items = item_groups[item_name]
                items.sort(key=lambda x: x['total_score'], reverse=True)
                # 只选每组第一个
                for i, item in enumerate(items):
                    item['selected'] = (i == 0)
                sorted_results.extend(items)
        
        # 添加未匹配的项目
        for item_name in unmatched_items:
            sorted_results.append({
                'item': item_name,
                'path': '',
                'filename': '未找到匹配文件',
                'item_score': 0,
                'keyword_bonus': 0,
                'total_score': 0,
                'matched_keywords': [],
                'selected': False,
                'matched': False  # 未匹配标记
            })
        
        self.search_results = sorted_results
        self.unmatched_count = len(unmatched_items)
        self.matched_count = len(matched_item_names)
        
        self.root.after(0, self.update_result_display)
        self.root.after(0, lambda: self.status_label.config(
            text=f"搜索完成: 项目总数 {len(item_names)}, 匹配 {self.matched_count}, 未匹配 {self.unmatched_count}"))
        self.root.after(0, lambda: self.progress.config(value=100))
    
    def update_result_display(self):
        """更新结果显示 - 仅对重复（多匹配）的项目名称着色"""
        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 统计每个项目的匹配数
        from collections import Counter
        item_counts = Counter(item['item'] for item in self.search_results)
        
        # 定义颜色方案 - 仅用于多匹配的项目（更深的颜色）
        duplicate_colors = [
            '#FFE4B5',  # 深一点的橙色
            '#E6E6FA',  # 深一点的紫色
            '#98FB98',  # 深一点的绿色
            '#FFB6C1',  # 深一点的粉色
            '#87CEEB',  # 深一点的蓝色
        ]
        
        # 只为有多个匹配的项目分配颜色
        duplicate_items = {item for item, count in item_counts.items() if count > 1}
        item_color_map = {}
        color_idx = 0
        
        for item in sorted(duplicate_items):
            item_color_map[item] = duplicate_colors[color_idx % len(duplicate_colors)]
            color_idx += 1
        
        # 配置标签样式
        for item, bg in item_color_map.items():
            tag_name = f'dup_{hash(item) % 10000}'
            self.tree.tag_configure(tag_name, background=bg)
        
        # 配置未匹配的样式
        self.tree.tag_configure('unmatched', background='#D9D9D9', foreground='#666666')
        
        # 填充数据
        for item in self.search_results:
            item_name = item['item']
            is_matched = item.get('matched', True)
            
            # 确定标签
            if not is_matched:
                # 未匹配的用灰色
                tag = 'unmatched'
            elif item_name in item_color_map:
                # 多匹配的用颜色
                tag = f'dup_{hash(item_name) % 10000}'
            else:
                # 单匹配的无颜色（默认）
                tag = ''
            
            self.tree.insert('', tk.END, values=(
                item['item'],
                item['filename'],
                f"{item['item_score']:.1f}" if item['item_score'] > 0 else '-',
                item['keyword_bonus'] if item['keyword_bonus'] > 0 else '-',
                f"{item['total_score']:.1f}" if item['total_score'] > 0 else '-',
                "是" if item['selected'] else "否"
            ), tags=(tag,))
    
    def toggle_copy(self, event):
        """双击切换是否复制"""
        item = self.tree.selection()[0]
        idx = self.tree.index(item)
        self.search_results[idx]['selected'] = not self.search_results[idx]['selected']
        self.tree.set(item, '是否复制', "是" if self.search_results[idx]['selected'] else "否")
    
    def select_all(self):
        """全选"""
        for item in self.search_results:
            item['selected'] = True
        self.update_result_display()
    
    def select_none(self):
        """全不选"""
        for item in self.search_results:
            item['selected'] = False
        self.update_result_display()
    
    def select_best(self):
        """仅选每组最高分"""
        # 按项目分组
        item_groups = defaultdict(list)
        for idx, item in enumerate(self.search_results):
            item_groups[item['item']].append(idx)
        
        # 先全不选
        for item in self.search_results:
            item['selected'] = False
        
        # 每组选第一个（已按评分排序）
        for item, indices in item_groups.items():
            if indices:
                self.search_results[indices[0]]['selected'] = True
        
        self.update_result_display()
    
    def export_excel(self):
        """导出Excel - 包含未匹配的项目"""
        target_dir = self.target_entry.get()
        if not target_dir:
            messagebox.showerror("错误", "请先设置目标目录")
            return
        
        os.makedirs(target_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = os.path.join(target_dir, f"搜索结果_{timestamp}.xlsx")
        
        # 生成Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "搜索结果"
        
        # 标题
        total_items = len(self.item_names) if hasattr(self, 'item_names') else len(self.search_results)
        matched_items = sum(1 for item in self.search_results if item.get('matched', True))
        unmatched_items = total_items - matched_items
        
        ws['A1'] = f"文件搜索结果 | 项目总数: {total_items} | 匹配: {matched_items} | 未匹配: {unmatched_items}"
        ws['A1'].font = Font(size=12, bold=True)
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表头
        headers = ['序号', '项目名称', '匹配数', '源文件路径', '源文件名', '目标文件名',
                   '名称匹配度', '关键词加分', '总分', '是否复制', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 按项目分组统计匹配数
        from collections import defaultdict
        item_groups = defaultdict(list)
        for item in self.search_results:
            if item.get('matched', True):
                item_groups[item['item']].append(item)
        
        # 数据
        row = 3
        for idx, item in enumerate(self.search_results, 1):
            is_matched = item.get('matched', True)
            match_count = len(item_groups.get(item['item'], [])) if is_matched else 0
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item['item'])
            ws.cell(row=row, column=3, value=match_count if is_matched else 0)
            ws.cell(row=row, column=4, value=item['path'] if is_matched else "未找到匹配文件")
            ws.cell(row=row, column=5, value=item['filename'] if is_matched else "-")
            ws.cell(row=row, column=6, value=item['filename'] if is_matched else "-")
            ws.cell(row=row, column=7, value=item['item_score'] if is_matched else "-")
            ws.cell(row=row, column=8, value=item['keyword_bonus'] if is_matched else "-")
            ws.cell(row=row, column=9, value=item['total_score'] if is_matched else "-")
            ws.cell(row=row, column=10, value="是" if item['selected'] else "否")
            ws.cell(row=row, column=11, value=", ".join(item['matched_keywords']) if is_matched else "未匹配")
            
            # 未匹配的项目用灰色背景
            if not is_matched:
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.font = Font(italic=True, color="666666")
                # 项目名称和匹配数列用深灰色突出
                ws.cell(row=row, column=2).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=3).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                ws.cell(row=row, column=3).font = Font(bold=True, color="FFFFFF")
            
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 45
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 30
        
        wb.save(excel_path)
        messagebox.showinfo("成功", f"Excel已导出到:\n{excel_path}")
    
    def start_copy(self):
        """开始复制"""
        target_dir = self.target_entry.get()
        if not target_dir:
            messagebox.showerror("错误", "请先设置目标目录")
            return
        
        selected_items = [item for item in self.search_results if item['selected']]
        if not selected_items:
            messagebox.showwarning("警告", "没有选择任何文件")
            return
        
        if not messagebox.askyesno("确认", f"确定要复制 {len(selected_items)} 个文件到 {target_dir} 吗？"):
            return
        
        # 在新线程中执行复制
        thread = threading.Thread(target=self.do_copy, args=(selected_items, target_dir))
        thread.daemon = True
        thread.start()
    
    def do_copy(self, items, target_dir):
        """执行复制"""
        os.makedirs(target_dir, exist_ok=True)
        
        copied = failed = 0
        for i, item in enumerate(items):
            progress = (i + 1) / len(items) * 100
            self.root.after(0, lambda p=progress: self.progress.config(value=p))
            self.root.after(0, lambda s=item['filename']: self.status_label.config(text=f"正在复制: {s}"))
            
            source = item['path']
            filename = item['filename']
            
            # 处理重名
            target_path = os.path.join(target_dir, filename)
            if os.path.exists(target_path):
                name, ext = os.path.splitext(filename)
                counter = 1
                while os.path.exists(target_path):
                    target_path = os.path.join(target_dir, f"{name}_{counter:03d}{ext}")
                    counter += 1
            
            try:
                shutil.copy2(source, target_path)
                copied += 1
            except Exception as e:
                failed += 1
                print(f"复制失败 {filename}: {e}")
        
        self.root.after(0, lambda: self.status_label.config(
            text=f"复制完成: 成功 {copied}, 失败 {failed}"))
        self.root.after(0, lambda: self.progress.config(value=100))
        self.root.after(0, lambda: messagebox.showinfo("完成", f"复制完成!\n成功: {copied}\n失败: {failed}"))


def main():
    root = tk.Tk()
    app = SearchApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
